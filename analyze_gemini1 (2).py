"""
analyze_gemini.py — Runs Gemini 2.0 Flash over every article in the labeling
sheet and writes structured analysis results back.

Setup:
  1. Get a free API key from https://aistudio.google.com/app/apikey
  2. export GEMINI_API_KEY="your_key_here"
  3. pip install google-genai openpyxl
  4. python analyze_gemini.py GKIN_Eval_Labeling.xlsx

Outputs:
  - New columns written to the articles sheet:
      gemini_manipulation_index, gemini_narrative_cluster,
      gemini_persuasion_count, gemini_claims_count
  - Full structured output saved to: gemini_results.json
  - Cached by article id so re-runs skip already-analyzed articles
"""

import os
import sys
import json
import time
import pathlib
from openpyxl import load_workbook
from google import genai
from google.genai import types

MODEL = "gemini-2.0-flash-001"
CACHE_PATH = "gemini_results.json"
SLEEP_BETWEEN = 2.0  # stay under free-tier rate limits

NARRATIVE_CLUSTERS = [
    "anti_vaccine", "election_fraud", "climate_denial", "immigration_threat",
    "economic_doom", "tech_surveillance", "elite_conspiracy", "pharma_corruption",
    "media_corruption", "cultural_decline", "historical_revisionism",
    "health_miracle", "financial_scam", "geopolitical_blame", "crime_panic",
    "gender_panic", "religious_persecution", "food_fear", "ai_panic", "none",
]

PROMPT_TEMPLATE = """You are a media literacy analyst. Given a news article, produce a structured JSON analysis identifying factual claims, persuasion techniques, emotional framing, missing context, and narrative patterns.

Analyze the article objectively. Do NOT take a political side. Focus on HOW the article is constructed to persuade, not WHETHER its conclusions are correct.

Return ONLY valid JSON matching this exact schema:

{{
  "claims": [
    {{
      "text": "direct quote or close paraphrase of the claim",
      "type": "Verifiable | Opinion | Unverifiable",
      "confidence": 0.0
    }}
  ],
  "persuasion_techniques": [
    {{
      "technique": "fear_appeal | bandwagon | loaded_language | false_urgency | authority_appeal | whataboutism | black_and_white",
      "span": "exact quote from article",
      "explanation": "one sentence on why this qualifies"
    }}
  ],
  "emotion_scores": {{
    "fear": 0,
    "anger": 0,
    "disgust": 0,
    "hope": 0,
    "guilt": 0,
    "ingroup_framing": 0
  }},
  "manipulation_index": 0,
  "missing_context": [
    {{
      "gap": "what context is missing",
      "why_it_matters": "one sentence on why this omission changes interpretation"
    }}
  ],
  "narrative_cluster": "one of the 20 categories or 'none'"
}}

Definitions:
- Verifiable: a factual claim checkable against public records, data, or authoritative sources.
- Opinion: a value judgment or interpretation.
- Unverifiable: a claim of fact that cannot practically be checked.
- fear_appeal: invoking threat or danger to motivate belief.
- bandwagon: appeal to popularity ("everyone thinks this").
- loaded_language: emotionally charged word choices where neutral alternatives exist.
- false_urgency: artificial time pressure.
- authority_appeal: citing authority instead of evidence.
- whataboutism: deflecting by pointing to unrelated wrongs.
- black_and_white: presenting complex issues as only two options.
- manipulation_index 0-20: straightforward reporting.
- manipulation_index 20-40: mild framing.
- manipulation_index 40-60: noticeable persuasive intent.
- manipulation_index 60-80: heavy manipulation.
- manipulation_index 80-100: blatant propaganda.

All emotion scores and manipulation_index are integers 0-100.
narrative_cluster must be one of: {clusters}.
Cap persuasion_techniques at 10. Cap claims at 15. Cap missing_context at 5.

Article:
\"\"\"
{article}
\"\"\"
"""


def find_sheet(wb):
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        if "url" in headers and "full_text" in headers:
            return ws, name
    return None, None


def ensure_columns(ws, new_cols):
    """Add new column headers if they don't exist. Returns dict name -> col index."""
    existing = {}
    for c in ws[1]:
        if c.value:
            existing[str(c.value).strip().lower()] = c.column
    next_col = max(existing.values()) + 1 if existing else 1
    for name in new_cols:
        if name not in existing:
            ws.cell(row=1, column=next_col, value=name)
            existing[name] = next_col
            next_col += 1
    return existing


def analyze_article(client, article_text: str) -> dict:
    prompt = PROMPT_TEMPLATE.format(
        clusters=", ".join(NARRATIVE_CLUSTERS),
        article=article_text[:12000],
    )
    response = client.models.generate_content(
        model=MODEL,
        contents=prompt,
        config=types.GenerateContentConfig(
            response_mime_type="application/json",
            temperature=0.2,
        ),
    )
    text = response.text
    if not text:
        raise ValueError("Empty response from Gemini")
    return json.loads(text)


def load_cache():
    if pathlib.Path(CACHE_PATH).exists():
        with open(CACHE_PATH) as f:
            return json.load(f)
    return {}


def save_cache(cache):
    with open(CACHE_PATH, "w") as f:
        json.dump(cache, f, indent=2)


def main(xlsx_path: str):
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("ERROR: GEMINI_API_KEY environment variable is not set.")
        print("Get a free key at https://aistudio.google.com/app/apikey")
        print('Then run:  export GEMINI_API_KEY="your_key_here"')
        sys.exit(1)

    client = genai.Client(api_key=api_key)

    wb = load_workbook(xlsx_path)
    ws, sheet_name = find_sheet(wb)
    if ws is None:
        print(f"No sheet with 'url' and 'full_text' columns. Sheets: {wb.sheetnames}")
        sys.exit(1)
    print(f"Using sheet: {sheet_name}")

    cols = ensure_columns(ws, [
        "gemini_manipulation_index",
        "gemini_narrative_cluster",
        "gemini_persuasion_count",
        "gemini_claims_count",
    ])
    col_id = cols.get("id")
    col_text = cols["full_text"]
    col_mi = cols["gemini_manipulation_index"]
    col_nc = cols["gemini_narrative_cluster"]
    col_pc = cols["gemini_persuasion_count"]
    col_cc = cols["gemini_claims_count"]

    cache = load_cache()
    ok = skipped = failed = 0

    for row in range(2, ws.max_row + 1):
        aid = ws.cell(row=row, column=col_id).value if col_id else row - 1
        text = ws.cell(row=row, column=col_text).value

        if not text or not isinstance(text, str) or len(text) < 200:
            continue
        if "FETCH FAILED" in text:
            continue

        key = str(aid)
        if key in cache:
            print(f"Article {aid}: cached")
            result = cache[key]
            skipped += 1
        else:
            print(f"Article {aid}: analyzing...")
            try:
                result = analyze_article(client, text)
                cache[key] = result
                save_cache(cache)
                ok += 1
                time.sleep(SLEEP_BETWEEN)
            except Exception as e:
                print(f"  ERROR: {e}")
                failed += 1
                continue

        # Write summary fields back to the sheet
        ws.cell(row=row, column=col_mi, value=result.get("manipulation_index"))
        ws.cell(row=row, column=col_nc, value=result.get("narrative_cluster"))
        ws.cell(row=row, column=col_pc, value=len(result.get("persuasion_techniques", [])))
        ws.cell(row=row, column=col_cc, value=len(result.get("claims", [])))

        mi = result.get("manipulation_index", "?")
        nc = result.get("narrative_cluster", "?")
        print(f"  -> manipulation={mi}  narrative={nc}")

    wb.save(xlsx_path)
    print("\n" + "=" * 50)
    print(f"Analyzed: {ok}  |  Cached: {skipped}  |  Failed: {failed}")
    print(f"Full results saved to: {CACHE_PATH}")
    print(f"Summary columns written to: {xlsx_path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python analyze_gemini.py <path_to_xlsx>")
        sys.exit(1)
    main(sys.argv[1])
