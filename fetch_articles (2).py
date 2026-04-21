"""
fetch_articles.py — Populates full_text (and title) in the eval labeling xlsx
from URLs already in the sheet.

Workflow:
  1. Paste URLs into the 'url' column of your labeling sheet.
  2. Run:  python fetch_articles.py GKIN_Eval_Labeling.xlsx
  3. Script fetches each URL, extracts clean article text, writes back.

Install deps once:
  pip install trafilatura openpyxl
"""

import sys
import time
import json
from openpyxl import load_workbook
import trafilatura

MAX_WORDS = 2000
SLEEP_BETWEEN = 1.0


def find_labels_sheet(wb):
    """Find the sheet that has a 'url' column in row 1."""
    for name in ["Labels", "labels", "Sheet1", "Sheet 1"]:
        if name in wb.sheetnames:
            ws = wb[name]
            if _has_url_header(ws):
                return ws, name
    for name in wb.sheetnames:
        ws = wb[name]
        if _has_url_header(ws):
            return ws, name
    return None, None


def _has_url_header(ws):
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == "url":
            return True
    return False


def get_column_indices(ws):
    cols = {}
    for cell in ws[1]:
        if cell.value:
            cols[str(cell.value).strip().lower()] = cell.column
    return cols


def fetch_article(url: str):
    try:
        downloaded = trafilatura.fetch_url(url)
        if not downloaded:
            return None, None
        result = trafilatura.extract(
            downloaded,
            output_format="json",
            with_metadata=True,
            include_comments=False,
            include_tables=False,
            favor_precision=True,
        )
        if not result:
            return None, None
        data = json.loads(result)
        title = (data.get("title") or "").strip()
        text = (data.get("text") or "").strip()
        words = text.split()
        if len(words) > MAX_WORDS:
            text = " ".join(words[:MAX_WORDS]) + " [...truncated]"
        return title, text
    except Exception as e:
        print(f"  ERROR: {e}")
        return None, None


def main(xlsx_path: str):
    wb = load_workbook(xlsx_path)
    ws, sheet_name = find_labels_sheet(wb)
    if ws is None:
        print("Could not find a sheet with a 'url' column.")
        print(f"Sheets in this file: {wb.sheetnames}")
        print("Make sure one of them has 'url' as a header in row 1.")
        sys.exit(1)
    print(f"Using sheet: {sheet_name}")

    cols = get_column_indices(ws)
    col_url = cols.get("url")
    col_title = cols.get("title")
    col_full_text = cols.get("full_text")

    if not col_url:
        print("No 'url' column found.")
        sys.exit(1)
    if not col_full_text:
        print("No 'full_text' column found. Add one to row 1.")
        sys.exit(1)

    updated = skipped = failed = 0

    for row in range(2, ws.max_row + 1):
        url = ws.cell(row=row, column=col_url).value
        existing_text = ws.cell(row=row, column=col_full_text).value

        if not url or not str(url).strip().startswith("http"):
            continue
        if existing_text and len(str(existing_text)) > 200:
            print(f"Row {row}: already has text, skipping")
            skipped += 1
            continue

        print(f"Row {row}: fetching {str(url)[:70]}...")
        title, text = fetch_article(str(url).strip())

        if not text:
            print("  FAILED")
            failed += 1
            ws.cell(row=row, column=col_full_text, value="[FETCH FAILED — paste manually]")
            continue

        ws.cell(row=row, column=col_full_text, value=text)
        if col_title and not ws.cell(row=row, column=col_title).value and title:
            ws.cell(row=row, column=col_title, value=title)

        print(f"  OK — {len(text.split())} words")
        updated += 1
        time.sleep(SLEEP_BETWEEN)

    wb.save(xlsx_path)
    print("\n" + "=" * 50)
    print(f"Updated: {updated}  |  Skipped: {skipped}  |  Failed: {failed}")
    print(f"Saved to {xlsx_path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python fetch_articles.py <path_to_xlsx>")
        sys.exit(1)
    main(sys.argv[1])