"""
discover_articles.py — Auto-populates url/title/category in the labeling sheet
by scraping a curated list of source pages.

Workflow:
  1. Run:  python discover_articles.py GKIN_Eval_Labeling.xlsx
     -> fills in url, title, category columns
  2. Then:  python fetch_articles.py GKIN_Eval_Labeling.xlsx
     -> fills in full_text column

Install deps once:
  pip install requests beautifulsoup4 openpyxl
"""

import sys
import time
import re
from urllib.parse import urljoin, urlparse
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# (source_url, category, count) — edit to taste
SOURCES = [
    # FACTUAL (10 total) — straight news, wire services, science
    ("https://apnews.com/hub/world-news", "factual", 3),
    ("https://www.aljazeera.com/news/", "factual", 3),
    ("https://www.npr.org/sections/world/", "factual", 2),
    ("https://arstechnica.com/science/", "factual", 2),
    # AMBIGUOUS (10 total) — opinion + advocacy from across the spectrum
    ("https://www.vox.com/", "ambiguous", 3),
    ("https://www.motherjones.com/politics/", "ambiguous", 3),
    ("https://www.nationalreview.com/", "ambiguous", 2),
    ("https://reason.com/", "ambiguous", 2),
    # MANIPULATIVE (10 total) — clickbait + health misinfo, balanced left/right
    ("https://www.naturalnews.com/", "manipulative", 4),
    ("https://www.palmerreport.com/", "manipulative", 3),
    ("https://www.thegatewaypundit.com/", "manipulative", 3),
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

EXCLUDE_SEGMENTS = {
    "tag", "tags", "category", "categories", "author", "authors",
    "about", "contact", "page", "pages", "search", "topics", "topic",
    "section", "sections", "video", "videos", "podcast", "podcasts",
    "subscribe", "newsletter", "newsletters", "shop", "store",
    "login", "signin", "account", "privacy", "terms", "rss", "feed",
    "hub", "live", "gallery", "photo", "photos",
}


def looks_like_article(href: str, base_domain: str) -> bool:
    if not href or href.startswith(("#", "mailto:", "javascript:")):
        return False
    parsed = urlparse(href)
    if parsed.netloc and base_domain not in parsed.netloc.replace("www.", ""):
        return False
    path = parsed.path
    segments = [s for s in path.split("/") if s]
    if len(segments) < 2:
        return False
    if any(s.lower() in EXCLUDE_SEGMENTS for s in segments):
        return False
    if re.fullmatch(r"\d{4}(/\d{1,2})?(/\d{1,2})?", path.strip("/")):
        return False
    # must contain a word (not just an id)
    if not re.search(r"[a-zA-Z]{5,}", segments[-1]):
        return False
    return True


def discover(source_url: str, count: int):
    try:
        r = requests.get(source_url, headers=HEADERS, timeout=20)
        r.raise_for_status()
    except Exception as e:
        print(f"  ERROR fetching {source_url}: {e}")
        return []
    soup = BeautifulSoup(r.text, "html.parser")
    base_domain = urlparse(source_url).netloc.replace("www.", "")
    seen = set()
    out = []
    for a in soup.find_all("a", href=True):
        href = urljoin(source_url, a["href"])
        href = href.split("#")[0].split("?")[0].rstrip("/")
        if href in seen:
            continue
        if not looks_like_article(href, base_domain):
            continue
        title = a.get_text(" ", strip=True)
        if len(title) < 20:  # too short to be a real headline
            continue
        seen.add(href)
        out.append((href, title))
        if len(out) >= count:
            break
    return out


def find_sheet(wb):
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        if "url" in headers and "category" in headers:
            return ws, name
    return None, None


def main(xlsx_path: str):
    wb = load_workbook(xlsx_path)
    ws, sheet_name = find_sheet(wb)
    if ws is None:
        print(f"No sheet with both 'url' and 'category' headers.")
        print(f"Sheets in file: {wb.sheetnames}")
        sys.exit(1)
    print(f"Using sheet: {sheet_name}")

    header_map = {str(c.value).strip().lower(): c.column for c in ws[1] if c.value}
    col_url = header_map["url"]
    col_title = header_map.get("title")
    col_cat = header_map["category"]

    all_articles = []
    for source_url, category, count in SOURCES:
        print(f"\n[{category}] {source_url}")
        found = discover(source_url, count)
        if not found:
            print("  (nothing found — site may be blocking or layout changed)")
        for url, title in found:
            all_articles.append((url, title, category))
            print(f"  + {title[:75]}")
        time.sleep(1.5)

    print(f"\nTotal discovered: {len(all_articles)} articles")

    # Write starting at row 2
    for i, (url, title, category) in enumerate(all_articles, start=2):
        ws.cell(row=i, column=col_url, value=url)
        if col_title:
            ws.cell(row=i, column=col_title, value=title)
        ws.cell(row=i, column=col_cat, value=category)

    wb.save(xlsx_path)
    print(f"\nSaved to {xlsx_path}")
    print("Next step: python fetch_articles.py " + xlsx_path)


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python discover_articles.py <path_to_xlsx>")
        sys.exit(1)
    main(sys.argv[1])
